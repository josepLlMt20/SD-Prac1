import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

EXCEL_FILE = "resultats_tests.xlsx"


def guardar_resultats(data_dicts, sheet_name):
    """
    Desa una llista de diccionaris com a DataFrame en un full específic dins del fitxer Excel.
    Si la pestanya ja existeix, afegeix les noves dades a sota. Si no, la crea.

    :param data_dicts: Llista de diccionaris amb resultats
    :param sheet_name: Nom del full (pestanya) dins del fitxer
    """
    df = pd.DataFrame(data_dicts)

    if os.path.exists(EXCEL_FILE):
        # Carregar el fitxer Excel existent
        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a') as writer:
            workbook = writer.book

            # Si la pestanya ja existeix, afegir les dades a sota
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                startrow = sheet.max_row  # Comença a escriure a la primera fila buida
                for r in dataframe_to_rows(df, index=False, header=False):
                    sheet.append(r)
            else:
                # Si la pestanya no existeix, crear-la i escriure les dades
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        # Si el fitxer no existeix, crear-lo i escriure les dades
        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
